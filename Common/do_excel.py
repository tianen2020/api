#!/usr/bin/env python
#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2020/7/4 21:31
# @Author  : 蒲天恩
# @File    : do_excel.py
from openpyxl import  load_workbook
from Common.do_path import excel_path

class Case:
    test_id = None
    model = None
    method = None
    url = None
    data = None
    expect = None

class DoExcel:
    def __init__(self,path_excel):
        self.wb = load_workbook(path_excel)

    def read_excel(self, sheet):
        sheet = self.wb[sheet]
        cases = []
        for i in range(2, sheet.max_row+1):
            case = Case()
            if sheet.cell(i, 1).value:
                case.test_id = sheet.cell(i, 1).value
                case.model = sheet.cell(i, 2).value
                case.method = sheet.cell(i, 3).value
                case.url = sheet.cell(i, 4).value
                case.data = sheet.cell(i, 5).value
                case.expect = sheet.cell(i, 6).value
                cases.append(case)
        self.wb.close()
        return cases

    def write_excel(self,sheet,row,column,new_value,path_excel):
        sheet = self.wb[sheet]
        sheet.cell(row, column).value = new_value
        self.wb.save(path_excel)
        self.wb.close()
# if __name__ == '__main__':
#     do = DoExcel(excel_path).read_excel("doctor")
#     for case in do:
#         print(case.data)




